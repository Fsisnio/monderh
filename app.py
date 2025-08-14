from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session, abort, Response, make_response
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_mail import Mail, Message
from flask_wtf import FlaskForm
from wtforms import StringField, TextAreaField, SelectField, FileField, DateField, TimeField, SubmitField, BooleanField, PasswordField
from wtforms.validators import DataRequired, Email, Length, EqualTo
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import os
from datetime import datetime, timedelta, timezone
import json
from functools import wraps
import io
from io import StringIO
import pickle
import csv
import pandas as pd
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from dotenv import load_dotenv

# Google API imports
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload

# Charger les variables d'environnement
load_dotenv()

app = Flask(__name__)
CORS(app)

# Configuration
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# Configuration de la base de données
DATABASE_URL = os.environ.get('DATABASE_URL')
if DATABASE_URL and DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL or 'sqlite:///monderh.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Email configuration
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', 'contact@monderh.fr')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', '')

# Google API configuration
app.config['GOOGLE_CLIENT_ID'] = os.environ.get('GOOGLE_CLIENT_ID', '')
app.config['GOOGLE_CLIENT_SECRET'] = os.environ.get('GOOGLE_CLIENT_SECRET', '')
app.config['GOOGLE_DISCOVERY_URL'] = "https://accounts.google.com/.well-known/openid_configuration"

# Initialize extensions
db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
mail = Mail(app)

# Create upload folder
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Database Models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    first_name = db.Column(db.String(50), nullable=False)
    last_name = db.Column(db.String(50), nullable=False)
    user_type = db.Column(db.String(20), default='candidate')  # candidate, client, admin
    company = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    is_active = db.Column(db.Boolean, default=True)
    
    # Relationships
    applications = db.relationship('Application', backref='applicant', lazy=True)
    appointments = db.relationship('Appointment', backref='user', lazy=True)
    
    def is_admin(self):
        return self.user_type == 'admin'

class Application(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    position = db.Column(db.String(100), nullable=False)
    service_type = db.Column(db.String(50), nullable=False)  # recrutement, coaching, formation, etc.
    cv_filename = db.Column(db.String(200))
    google_drive_link = db.Column(db.String(500))
    cover_letter = db.Column(db.Text)
    linkedin_url = db.Column(db.String(200))
    experience_years = db.Column(db.Integer)
    salary_expectation = db.Column(db.String(50))
    availability = db.Column(db.String(50))
    status = db.Column(db.String(20), default='pending')  # pending, reviewed, accepted, rejected
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    notes = db.Column(db.Text)

class Appointment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    service_type = db.Column(db.String(50), nullable=False)
    date = db.Column(db.Date, nullable=False)
    time = db.Column(db.Time, nullable=False)
    duration = db.Column(db.Integer, default=60)  # minutes
    subject = db.Column(db.String(200))
    description = db.Column(db.Text)
    google_calendar_link = db.Column(db.String(500))
    status = db.Column(db.String(20), default='pending')  # pending, confirmed, cancelled, completed
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))

class JobOffer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    company = db.Column(db.String(100), nullable=False)
    location = db.Column(db.String(100), nullable=False)
    contract_type = db.Column(db.String(50), nullable=False)  # CDI, CDD, Stage, Alternance
    experience_level = db.Column(db.String(50), nullable=False)  # Junior, Confirmé, Senior, Expert
    salary_range = db.Column(db.String(100))
    description = db.Column(db.Text, nullable=False)
    requirements = db.Column(db.Text, nullable=False)
    benefits = db.Column(db.Text)
    department = db.Column(db.String(100))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))
    
    # Relationships
    applications = db.relationship('JobApplication', backref='job_offer', lazy=True)
    
    def __repr__(self):
        return f'<JobOffer {self.title} at {self.company}>'

class JobApplication(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    job_offer_id = db.Column(db.Integer, db.ForeignKey('job_offer.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    cv_filename = db.Column(db.String(200))
    cover_letter = db.Column(db.Text)
    status = db.Column(db.String(20), default='pending')  # pending, reviewed, accepted, rejected
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    notes = db.Column(db.Text)

class Newsletter(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    first_name = db.Column(db.String(50))
    last_name = db.Column(db.String(50))
    company = db.Column(db.String(100))
    interests = db.Column(db.Text)  # JSON string of service interests
    subscribed_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    is_active = db.Column(db.Boolean, default=True)

class SiteSettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    site_name = db.Column(db.String(100), default='MondeRH')
    site_description = db.Column(db.Text, default='Votre partenaire en ressources humaines')
    contact_email = db.Column(db.String(120), default='contact@monderh.fr')
    contact_phone = db.Column(db.String(20), default='+33 1 23 45 67 89')
    address = db.Column(db.Text, default='123 Avenue des Ressources Humaines, 75001 Paris')
    logo_url = db.Column(db.String(200))
    hero_title = db.Column(db.String(200), default='Trouvez votre carrière idéale')
    hero_subtitle = db.Column(db.Text, default='Nous vous accompagnons dans votre parcours professionnel')
    facebook_url = db.Column(db.String(200))
    linkedin_url = db.Column(db.String(200))
    twitter_url = db.Column(db.String(200))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))

class GoogleToken(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    access_token = db.Column(db.Text)
    refresh_token = db.Column(db.Text)
    token_uri = db.Column(db.String(500))
    client_id = db.Column(db.String(500))
    client_secret = db.Column(db.String(500))
    scopes = db.Column(db.Text)  # JSON array of scopes
    expiry = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))
    
    user = db.relationship('User', backref=db.backref('google_tokens', lazy=True))

# Admin decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin():
            abort(403)
        return f(*args, **kwargs)
    return decorated_function

# Google API utilities
def get_google_credentials(user_id):
    """Récupère les credentials Google pour un utilisateur"""
    token = GoogleToken.query.filter_by(user_id=user_id).first()
    if not token:
        return None
    
    creds_info = {
        'token': token.access_token,
        'refresh_token': token.refresh_token,
        'token_uri': token.token_uri,
        'client_id': token.client_id,
        'client_secret': token.client_secret,
        'scopes': json.loads(token.scopes) if token.scopes else []
    }
    
    if token.expiry:
        creds_info['expiry'] = token.expiry.isoformat()
    
    creds = Credentials.from_authorized_user_info(creds_info)
    
    # Vérifier si le token doit être rafraîchi
    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
        # Sauvegarder le nouveau token
        save_google_credentials(user_id, creds)
    
    return creds

def save_google_credentials(user_id, credentials):
    """Sauvegarde les credentials Google d'un utilisateur"""
    token = GoogleToken.query.filter_by(user_id=user_id).first()
    if not token:
        token = GoogleToken(user_id=user_id)
    
    token.access_token = credentials.token
    token.refresh_token = credentials.refresh_token
    token.token_uri = credentials.token_uri
    token.client_id = credentials.client_id
    token.client_secret = credentials.client_secret
    token.scopes = json.dumps(credentials.scopes)
    if credentials.expiry:
        token.expiry = credentials.expiry.replace(tzinfo=None)
    
    db.session.add(token)
    db.session.commit()

def upload_to_google_drive(user_id, file_content, filename, mimetype):
    """Upload un fichier vers Google Drive"""
    creds = get_google_credentials(user_id)
    if not creds:
        return None
    
    try:
        service = build('drive', 'v3', credentials=creds)
        
        # Créer le dossier MondeRH s'il n'existe pas
        folder_name = 'MondeRH_CV'
        folders = service.files().list(
            q=f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder'",
            fields="files(id, name)"
        ).execute()
        
        if not folders.get('files'):
            folder_metadata = {
                'name': folder_name,
                'mimeType': 'application/vnd.google-apps.folder'
            }
            folder = service.files().create(body=folder_metadata, fields='id').execute()
            folder_id = folder.get('id')
        else:
            folder_id = folders.get('files')[0].get('id')
        
        # Upload le fichier
        file_metadata = {
            'name': filename,
            'parents': [folder_id]
        }
        
        media = MediaIoBaseUpload(io.BytesIO(file_content), mimetype=mimetype)
        file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields='id,webViewLink'
        ).execute()
        
        return {
            'file_id': file.get('id'),
            'web_link': file.get('webViewLink')
        }
        
    except Exception as e:
        print(f"Erreur upload Google Drive: {e}")
        return None

def create_calendar_event(user_id, appointment_data):
    """Crée un événement dans Google Calendar"""
    creds = get_google_credentials(user_id)
    if not creds:
        return None
    
    try:
        service = build('calendar', 'v3', credentials=creds)
        
        # Convertir la date et l'heure en format ISO
        start_datetime = datetime.combine(appointment_data['date'], appointment_data['time'])
        end_datetime = start_datetime + timedelta(minutes=appointment_data.get('duration', 60))
        
        event = {
            'summary': f"RDV {appointment_data['service_type']} - MondeRH",
            'description': f"Rendez-vous {appointment_data['service_type']}\n"
                          f"Client: {appointment_data['client_name']}\n"
                          f"Sujet: {appointment_data.get('subject', 'Non spécifié')}\n"
                          f"Description: {appointment_data.get('description', '')}",
            'start': {
                'dateTime': start_datetime.isoformat(),
                'timeZone': 'Europe/Paris',
            },
            'end': {
                'dateTime': end_datetime.isoformat(),
                'timeZone': 'Europe/Paris',
            },
            'attendees': [
                {'email': appointment_data.get('client_email', '')}
            ],
            'reminders': {
                'useDefault': False,
                'overrides': [
                    {'method': 'email', 'minutes': 24 * 60},
                    {'method': 'popup', 'minutes': 30},
                ],
            },
        }
        
        event = service.events().insert(calendarId='primary', body=event).execute()
        return event.get('htmlLink')
        
    except Exception as e:
        print(f"Erreur création événement Calendar: {e}")
        return None

# Forms
class LoginForm(FlaskForm):
    email = StringField('Email', validators=[DataRequired(), Email()])
    password = PasswordField('Mot de passe', validators=[DataRequired()])
    remember_me = BooleanField('Se souvenir de moi')
    submit = SubmitField('Se connecter')

class RegistrationForm(FlaskForm):
    first_name = StringField('Prénom', validators=[DataRequired()])
    last_name = StringField('Nom', validators=[DataRequired()])
    email = StringField('Email', validators=[DataRequired(), Email()])
    account_type = SelectField('Type de compte', choices=[
        ('candidate', 'Candidat'),
        ('client', 'Client'),
        ('admin', 'Administrateur')
    ])
    password = PasswordField('Mot de passe', validators=[DataRequired(), Length(min=6)])
    confirm_password = PasswordField('Confirmer le mot de passe', validators=[
        DataRequired(), 
        EqualTo('password', message='Les mots de passe doivent correspondre')
    ])
    accept_terms = BooleanField('J\'accepte les conditions d\'utilisation', validators=[DataRequired()])
    company = StringField('Entreprise')
    phone = StringField('Téléphone')
    submit = SubmitField('S\'inscrire')

class ApplicationForm(FlaskForm):
    position = StringField('Poste recherché', validators=[DataRequired()])
    service_type = SelectField('Type de service', choices=[
        ('recrutement', 'Recrutement'),
        ('coaching', 'Coaching'),
        ('formation', 'Formation'),
        ('interim', 'Intérim'),
        ('conseil', 'Conseil en Organisation')
    ])
    experience_years = SelectField('Années d\'expérience', choices=[
        ('0-1', '0-1 an'),
        ('1-3', '1-3 ans'),
        ('3-5', '3-5 ans'),
        ('5-10', '5-10 ans'),
        ('10+', '10+ ans')
    ])
    salary_expectation = StringField('Prétentions salariales')
    availability = SelectField('Disponibilité', choices=[
        ('immediate', 'Immédiate'),
        ('1-2-weeks', '1-2 semaines'),
        ('1-month', '1 mois'),
        ('3-months', '3 mois'),
        ('flexible', 'Flexible')
    ])

    cover_letter = TextAreaField('Lettre de motivation')
    cv_file = FileField('CV (PDF)')
    submit = SubmitField('Soumettre ma candidature')

class AppointmentForm(FlaskForm):
    service_type = SelectField('Service', choices=[
        ('recrutement', 'Recrutement'),
        ('coaching', 'Coaching'),
        ('formation', 'Formation'),
        ('interim', 'Intérim'),
        ('conseil', 'Conseil en Organisation')
    ], validators=[DataRequired()])
    date = DateField('Date', validators=[DataRequired()])
    time = TimeField('Heure', validators=[DataRequired()])
    duration = SelectField('Durée', choices=[
        (30, '30 minutes'),
        (60, '1 heure'),
        (90, '1h30'),
        (120, '2 heures')
    ], default=60)
    subject = StringField('Sujet', validators=[DataRequired()])
    description = TextAreaField('Description')
    submit = SubmitField('Prendre rendez-vous')

class NewsletterForm(FlaskForm):
    email = StringField('Email', validators=[DataRequired(), Email()])
    first_name = StringField('Prénom')
    last_name = StringField('Nom')
    company = StringField('Entreprise')
    interests = SelectField('Services d\'intérêt', choices=[
        ('all', 'Tous les services'),
        ('recrutement', 'Recrutement'),
        ('coaching', 'Coaching'),
        ('formation', 'Formation'),
        ('interim', 'Intérim'),
        ('conseil', 'Conseil en Organisation')
    ])
    submit = SubmitField('S\'abonner')

class SiteSettingsForm(FlaskForm):
    site_name = StringField('Nom du site', validators=[DataRequired()])
    site_description = TextAreaField('Description du site')
    contact_email = StringField('Email de contact', validators=[DataRequired(), Email()])
    contact_phone = StringField('Téléphone de contact')
    address = TextAreaField('Adresse')
    hero_title = StringField('Titre principal de la page d\'accueil')
    hero_subtitle = TextAreaField('Sous-titre de la page d\'accueil')
    logo_url = StringField('URL du Logo')
    facebook_url = StringField('URL Facebook')
    linkedin_url = StringField('URL LinkedIn')
    twitter_url = StringField('URL Twitter')
    submit = SubmitField('Enregistrer les paramètres')

class JobOfferForm(FlaskForm):
    title = StringField('Titre du poste', validators=[DataRequired(), Length(min=5, max=200)])
    company = StringField('Entreprise', validators=[DataRequired(), Length(min=2, max=100)])
    location = StringField('Localisation', validators=[DataRequired(), Length(min=2, max=100)])
    contract_type = SelectField('Type de contrat', choices=[
        ('CDI', 'CDI'),
        ('CDD', 'CDD'),
        ('Stage', 'Stage'),
        ('Alternance', 'Alternance'),
        ('Freelance', 'Freelance'),
        ('Intérim', 'Intérim')
    ], validators=[DataRequired()])
    experience_level = SelectField('Niveau d\'expérience', choices=[
        ('Junior', 'Junior (0-2 ans)'),
        ('Confirmé', 'Confirmé (3-5 ans)'),
        ('Senior', 'Senior (6-10 ans)'),
        ('Expert', 'Expert (10+ ans)')
    ], validators=[DataRequired()])
    salary_range = StringField('Fourchette salariale (ex: 30000-45000€)')
    description = TextAreaField('Description du poste', validators=[DataRequired(), Length(min=50)])
    requirements = TextAreaField('Profil recherché', validators=[DataRequired(), Length(min=30)])
    benefits = TextAreaField('Avantages et bénéfices')
    department = StringField('Département')
    is_active = BooleanField('Offre active')
    submit = SubmitField('Publier l\'offre')

class UserEditForm(FlaskForm):
    first_name = StringField('Prénom', validators=[DataRequired()])
    last_name = StringField('Nom', validators=[DataRequired()])
    email = StringField('Email', validators=[DataRequired(), Email()])
    user_type = SelectField('Type d\'utilisateur', choices=[
        ('candidate', 'Candidat'),
        ('client', 'Client'),
        ('admin', 'Administrateur')
    ])
    company = StringField('Entreprise')
    phone = StringField('Téléphone')
    is_active = BooleanField('Compte actif')
    submit = SubmitField('Sauvegarder')

# User loader for Flask-Login
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Filtres Jinja2 personnalisés
@app.template_filter('nl2br')
def nl2br(value):
    """Convertit les retours à la ligne en balises <br> et nettoie le contenu"""
    if value:
        import html
        # Décoder les entités HTML
        decoded = html.unescape(value)
        # Nettoyer les guillemets indésirables et caractères spéciaux
        cleaned = decoded.replace('"', '').replace('"', '').replace('"', '').replace(''', "'").replace(''', "'")
        # Convertir les retours à la ligne en balises <br>
        return cleaned.replace('\n', '<br>')
    return value

# Données des services (inchangées)
services = {
    'recrutement': {
        'title': 'Recrutement',
        'description': 'Nous accompagnons les entreprises dans leur processus de recrutement pour trouver les meilleurs talents.',
        'features': [
            'Analyse des besoins en recrutement',
            'Sourcing et présélection de candidats',
            'Entretiens et évaluation des compétences',
            'Accompagnement dans la prise de décision',
            'Intégration des nouveaux collaborateurs'
        ],
        'icon': '👥'
    },
    'coaching': {
        'title': 'Coaching',
        'description': 'Développez votre potentiel et celui de vos équipes grâce à nos programmes de coaching personnalisés.',
        'features': [
            'Coaching individuel et d\'équipe',
            'Développement du leadership',
            'Gestion du stress et de la performance',
            'Accompagnement au changement',
            'Optimisation des relations professionnelles'
        ],
        'icon': '🎯'
    },
    'formation': {
        'title': 'Formation',
        'description': 'Formez vos équipes avec nos programmes sur mesure adaptés à vos enjeux business.',
        'features': [
            'Formations sur mesure',
            'Développement des compétences managériales',
            'Formation aux outils RH',
            'E-learning et blended learning',
            'Certifications professionnelles'
        ],
        'icon': '📚'
    },
    'interim': {
        'title': 'Intérim',
        'description': 'Solutions d\'intérim adaptées pour répondre à vos besoins temporaires en personnel qualifié.',
        'features': [
            'Mise à disposition de personnel qualifié',
            'Gestion administrative complète',
            'Flexibilité et réactivité',
            'Profils adaptés à vos besoins',
            'Suivi et accompagnement'
        ],
        'icon': '⚡'
    },
    'conseil': {
        'title': 'Conseil en Organisation',
        'description': 'Optimisez votre organisation et vos processus RH pour améliorer la performance de votre entreprise.',
        'features': [
            'Audit organisationnel',
            'Restructuration et réorganisation',
            'Optimisation des processus RH',
            'Accompagnement au changement',
            'Pilotage de la performance'
        ],
        'icon': '🏢'
    }
}

# Routes principales
@app.route('/')
def index():
    newsletter_form = NewsletterForm()
    return render_template('index.html', services=services, newsletter_form=newsletter_form)

@app.route('/api/services')
def get_services():
    return jsonify(services)

@app.route('/service/<service_name>')
def service_detail(service_name):
    if service_name not in services:
        return render_template('404.html'), 404
    
    # Utiliser la page améliorée pour le recrutement
    if service_name == 'recrutement':
        return render_template('recruitment_enhanced.html')
    
    # Utiliser la page améliorée pour le coaching
    if service_name == 'coaching':
        return render_template('coaching_enhanced.html')
    
    # Utiliser la page améliorée pour la formation
    if service_name == 'formation':
        return render_template('formation_enhanced.html')
    
    # Utiliser la page améliorée pour l'intérim
    if service_name == 'interim':
        return render_template('interim_enhanced.html')
    
    # Utiliser la page améliorée pour le conseil
    if service_name == 'conseil':
        return render_template('conseil_enhanced.html')
    
    return render_template('service_detail.html', service=services[service_name], service_name=service_name)

# Authentication routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user and check_password_hash(user.password_hash, form.password.data):
            login_user(user, remember=form.remember_me.data)
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        else:
            flash('Email ou mot de passe incorrect', 'error')
    
    return render_template('auth/login.html', form=form)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = RegistrationForm()
    if form.validate_on_submit():
        if User.query.filter_by(email=form.email.data).first():
            flash('Cet email est déjà utilisé', 'error')
        else:
            user = User(
                email=form.email.data,
                password_hash=generate_password_hash(form.password.data),
                first_name=form.first_name.data,
                last_name=form.last_name.data,
                user_type=form.account_type.data,
                company=form.company.data,
                phone=form.phone.data
            )
            db.session.add(user)
            db.session.commit()
            flash('Compte créé avec succès ! Vous pouvez maintenant vous connecter.', 'success')
            return redirect(url_for('login'))
    
    return render_template('auth/register.html', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

# Dashboard routes
@app.route('/dashboard')
@login_required
def dashboard():
    user_applications = Application.query.filter_by(user_id=current_user.id).order_by(Application.created_at.desc()).all()
    user_appointments = Appointment.query.filter_by(user_id=current_user.id).order_by(Appointment.date.desc()).all()
    
    # Vérifier le statut Google seulement pour les administrateurs
    google_connected = False
    if current_user.is_admin():
        google_token = GoogleToken.query.filter_by(user_id=current_user.id).first()
        google_connected = google_token is not None
    
    return render_template('dashboard/dashboard.html', 
                         applications=user_applications, 
                         appointments=user_appointments,
                         google_connected=google_connected)

# Application routes
@app.route('/apply', methods=['GET', 'POST'])
def apply():
    form = ApplicationForm()
    if form.validate_on_submit():
        # Handle file upload
        cv_filename = None
        google_drive_link = None
        
        if form.cv_file.data:
            file = form.cv_file.data
            if file and allowed_file(file.filename):
                filename = secure_filename(f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
                
                # Sauvegarde locale (backup)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                cv_filename = filename
                
                # Upload vers Google Drive si l'utilisateur est administrateur et a autorisé Google
                if current_user.is_authenticated and current_user.is_admin():
                    file.seek(0)  # Reset file pointer
                    file_content = file.read()
                    drive_result = upload_to_google_drive(
                        current_user.id, 
                        file_content, 
                        filename, 
                        file.mimetype
                    )
                    if drive_result:
                        google_drive_link = drive_result['web_link']
                        flash('CV sauvegardé sur Google Drive !', 'info')
        
        application = Application(
            user_id=current_user.id if current_user.is_authenticated else None,
            position=form.position.data,
            service_type=form.service_type.data,
            cv_filename=cv_filename,
            google_drive_link=google_drive_link,
            cover_letter=form.cover_letter.data,

            experience_years=form.experience_years.data,
            salary_expectation=form.salary_expectation.data,
            availability=form.availability.data
        )
        
        db.session.add(application)
        db.session.commit()
        
        # Send notification email
        send_application_notification(application)
        
        flash('Votre candidature a été soumise avec succès !', 'success')
        return redirect(url_for('dashboard') if current_user.is_authenticated else url_for('index'))
    
    return render_template('apply.html', form=form)

# Appointment routes
@app.route('/appointments', methods=['GET', 'POST'])
@login_required
def appointments():
    form = AppointmentForm()
    if form.validate_on_submit():
        # Créer l'événement Google Calendar seulement pour les administrateurs
        calendar_link = None
        if current_user.is_admin():
            appointment_data = {
                'service_type': form.service_type.data,
                'date': form.date.data,
                'time': form.time.data,
                'duration': form.duration.data,
                'subject': form.subject.data,
                'description': form.description.data,
                'client_name': f"{current_user.first_name} {current_user.last_name}",
                'client_email': current_user.email
            }
            
            calendar_link = create_calendar_event(current_user.id, appointment_data)
        
        appointment = Appointment(
            user_id=current_user.id,
            service_type=form.service_type.data,
            date=form.date.data,
            time=form.time.data,
            duration=form.duration.data,
            subject=form.subject.data,
            description=form.description.data,
            google_calendar_link=calendar_link
        )
        
        db.session.add(appointment)
        db.session.commit()
        
        # Send confirmation email
        send_appointment_confirmation(appointment)
        
        if calendar_link:
            flash('Rendez-vous ajouté à votre Google Calendar !', 'info')
        
        flash('Rendez-vous pris avec succès ! Vous recevrez une confirmation par email.', 'success')
        return redirect(url_for('dashboard'))
    
    user_appointments = Appointment.query.filter_by(user_id=current_user.id).order_by(Appointment.date.desc()).all()
    return render_template('appointments.html', form=form, appointments=user_appointments)

# Newsletter routes
@app.route('/newsletter', methods=['POST'])
def subscribe_newsletter():
    form = NewsletterForm()
    if form.validate_on_submit():
        # Check if already subscribed
        existing = Newsletter.query.filter_by(email=form.email.data).first()
        if existing:
            flash('Cet email est déjà inscrit à notre newsletter.', 'info')
        else:
            newsletter = Newsletter(
                email=form.email.data,
                first_name=form.first_name.data,
                last_name=form.last_name.data,
                company=form.company.data,
                interests=form.interests.data
            )
            db.session.add(newsletter)
            db.session.commit()
            flash('Inscription à la newsletter réussie !', 'success')
    
    return redirect(request.referrer or url_for('index'))

# API routes
@app.route('/api/contact', methods=['POST'])
def contact():
    data = request.get_json()
    # Send contact email
    send_contact_email(data)
    return jsonify({'message': 'Message reçu avec succès! Nous vous recontacterons rapidement.'})



# Chat functionality
@app.route('/api/chat/message', methods=['POST'])
def chat_message():
    data = request.get_json()
    # Here you would integrate with a chat service
    # For now, we'll return a simple response
    return jsonify({
        'message': 'Merci pour votre message. Un consultant vous répondra dans les plus brefs délais.',
        'timestamp': datetime.now().isoformat()
    })

# Utility functions
def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def send_application_notification(application):
    try:
        msg = Message(
            'Nouvelle candidature reçue - MondeRH',
            sender=app.config['MAIL_USERNAME'],
            recipients=[app.config['MAIL_USERNAME']]
        )
        msg.body = f"""
        Nouvelle candidature reçue :
        
        Poste : {application.position}
        Service : {application.service_type}
        Candidat : {application.applicant.first_name} {application.applicant.last_name}
        Email : {application.applicant.email}
        
        Consultez le dashboard pour plus de détails.
        """
        mail.send(msg)
    except Exception as e:
        print(f"Error sending email: {e}")

def send_appointment_confirmation(appointment):
    try:
        msg = Message(
            'Confirmation de rendez-vous - MondeRH',
            sender=app.config['MAIL_USERNAME'],
            recipients=[appointment.user.email]
        )
        msg.body = f"""
        Bonjour {appointment.user.first_name},
        
        Votre rendez-vous a été confirmé :
        
        Service : {appointment.service_type}
        Date : {appointment.date.strftime('%d/%m/%Y')}
        Heure : {appointment.time.strftime('%H:%M')}
        Durée : {appointment.duration} minutes
        Sujet : {appointment.subject}
        
        Nous vous contacterons pour confirmer les détails.
        
        Cordialement,
        L'équipe MondeRH
        """
        mail.send(msg)
    except Exception as e:
        print(f"Error sending email: {e}")

def send_contact_email(data):
    try:
        msg = Message(
            'Nouveau message de contact - MondeRH',
            sender=app.config['MAIL_USERNAME'],
            recipients=[app.config['MAIL_USERNAME']]
        )
        msg.body = f"""
        Nouveau message de contact :
        
        Nom : {data.get('firstName', '')} {data.get('lastName', '')}
        Email : {data.get('email', '')}
        Entreprise : {data.get('company', '')}
        Service : {data.get('service', '')}
        Message : {data.get('message', '')}
        """
        mail.send(msg)
    except Exception as e:
        print(f"Error sending email: {e}")

def send_application_accepted_email(application):
    """Envoyer un email de notification d'acceptation de candidature"""
    try:
        msg = Message(
            'Candidature acceptée - MondeRH',
            sender=app.config['MAIL_USERNAME'],
            recipients=[application.applicant.email]
        )
        msg.body = f"""
        Bonjour {application.applicant.first_name},
        
        Nous avons le plaisir de vous informer que votre candidature pour le poste de "{application.position}" a été acceptée !
        
        Détails de votre candidature :
        - Poste : {application.position}
        - Service : {application.service_type}
        - Expérience : {application.experience_years} ans
        
        Notre équipe vous contactera dans les plus brefs délais pour organiser la suite du processus.
        
        Félicitations !
        
        Cordialement,
        L'équipe MondeRH
        """
        mail.send(msg)
    except Exception as e:
        print(f"Error sending acceptance email: {e}")

def send_application_rejected_email(application):
    """Envoyer un email de notification de rejet de candidature"""
    try:
        msg = Message(
            'Réponse à votre candidature - MondeRH',
            sender=app.config['MAIL_USERNAME'],
            recipients=[application.applicant.email]
        )
        msg.body = f"""
        Bonjour {application.applicant.first_name},
        
        Nous vous remercions pour votre candidature pour le poste de "{application.position}".
        
        Après avoir étudié attentivement votre profil, nous regrettons de vous informer que nous ne pouvons pas retenir votre candidature pour ce poste.
        
        Nous vous encourageons à continuer à nous suivre pour d'autres opportunités qui pourraient correspondre à votre profil.
        
        Nous vous souhaitons le meilleur pour la suite de votre carrière.
        
        Cordialement,
        L'équipe MondeRH
        """
        mail.send(msg)
    except Exception as e:
        print(f"Error sending rejection email: {e}")

def send_application_reviewed_email(application):
    """Envoyer un email de notification de révision de candidature"""
    try:
        msg = Message(
            'Candidature en cours d\'examen - MondeRH',
            sender=app.config['MAIL_USERNAME'],
            recipients=[application.applicant.email]
        )
        msg.body = f"""
        Bonjour {application.applicant.first_name},
        
        Nous vous informons que votre candidature pour le poste de "{application.position}" est actuellement en cours d'examen par notre équipe.
        
        Détails de votre candidature :
        - Poste : {application.position}
        - Service : {application.service_type}
        - Expérience : {application.experience_years} ans
        
        Nous vous tiendrons informé(e) dès qu'une décision sera prise.
        
        Merci pour votre patience.
        
        Cordialement,
        L'équipe MondeRH
        """
        mail.send(msg)
    except Exception as e:
        print(f"Error sending review email: {e}")

# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('500.html'), 500

# Admin routes
@app.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    # Statistiques générales
    total_users = User.query.count()
    total_applications = Application.query.count()
    total_appointments = Appointment.query.count()
    total_newsletters = Newsletter.query.filter_by(is_active=True).count()
    
    # Données récentes
    recent_applications = Application.query.order_by(Application.created_at.desc()).limit(5).all()
    recent_appointments = Appointment.query.order_by(Appointment.created_at.desc()).limit(5).all()
    
    return render_template('admin/dashboard.html',
                         total_users=total_users,
                         total_applications=total_applications,
                         total_appointments=total_appointments,
                         total_newsletters=total_newsletters,
                         recent_applications=recent_applications,
                         recent_appointments=recent_appointments)

@app.route('/admin/export/<format>')
@login_required
@admin_required
def admin_export(format):
    """Exporte les données du tableau de bord dans différents formats"""
    if format == 'csv':
        return export_csv()
    elif format == 'excel':
        return export_excel()
    elif format == 'pdf':
        return export_pdf()
    elif format == 'png':
        return export_png()
    else:
        flash('Format d\'export non supporté', 'error')
        return redirect(url_for('admin_dashboard'))

def export_csv():
    """Exporte les données en CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # En-têtes
    writer.writerow(['Rapport MondeRH - ' + datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')])
    writer.writerow([])
    
    # Statistiques générales
    writer.writerow(['STATISTIQUES GÉNÉRALES'])
    writer.writerow(['Métrique', 'Valeur'])
    writer.writerow(['Total Utilisateurs', User.query.count()])
    writer.writerow(['Total Candidatures', Application.query.count()])
    writer.writerow(['Total Rendez-vous', Appointment.query.count()])
    writer.writerow(['Newsletters Actives', Newsletter.query.filter_by(is_active=True).count()])
    writer.writerow(['Offres d\'emploi', JobOffer.query.count()])
    writer.writerow(['Offres actives', JobOffer.query.filter_by(is_active=True).count()])
    writer.writerow([])
    
    # Candidatures récentes
    writer.writerow(['CANDIDATURES RÉCENTES'])
    writer.writerow(['ID', 'Candidat', 'Poste', 'Service', 'Statut', 'Date'])
    recent_applications = Application.query.order_by(Application.created_at.desc()).limit(10).all()
    for app in recent_applications:
        writer.writerow([
            app.id,
            f"{app.applicant.first_name} {app.applicant.last_name}",
            app.position,
            app.service_type,
            app.status,
            app.created_at.strftime('%d/%m/%Y %H:%M')
        ])
    writer.writerow([])
    
    # Rendez-vous récents
    writer.writerow(['RENDEZ-VOUS RÉCENTS'])
    writer.writerow(['ID', 'Client', 'Service', 'Date', 'Heure', 'Statut'])
    recent_appointments = Appointment.query.order_by(Appointment.created_at.desc()).limit(10).all()
    for apt in recent_appointments:
        writer.writerow([
            apt.id,
            f"{apt.user.first_name} {apt.user.last_name}",
            apt.service_type,
            apt.date.strftime('%d/%m/%Y'),
            apt.time.strftime('%H:%M'),
            apt.status
        ])
    writer.writerow([])
    
    # Offres d'emploi
    writer.writerow(['OFFRES D\'EMPLOI'])
    writer.writerow(['ID', 'Titre', 'Entreprise', 'Localisation', 'Type', 'Statut', 'Candidatures'])
    job_offers = JobOffer.query.all()
    for job in job_offers:
        writer.writerow([
            job.id,
            job.title,
            job.company,
            job.location,
            job.contract_type,
            'Active' if job.is_active else 'Inactive',
            len(job.applications)
        ])
    
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename=rapport_monderh_{datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")}.csv'}
    )

def export_excel():
    """Exporte les données en Excel"""
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    wb.remove(wb.active)
    
    # Statistiques générales
    ws_stats = wb.create_sheet("Statistiques")
    ws_stats.title = "Statistiques"
    
    # Titre
    ws_stats['A1'] = f"Rapport MondeRH - {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')}"
    ws_stats['A1'].font = Font(size=16, bold=True)
    ws_stats.merge_cells('A1:C1')
    
    # Statistiques
    stats_data = [
        ['Métrique', 'Valeur'],
        ['Total Utilisateurs', User.query.count()],
        ['Total Candidatures', Application.query.count()],
        ['Total Rendez-vous', Appointment.query.count()],
        ['Newsletters Actives', Newsletter.query.filter_by(is_active=True).count()],
        ['Offres d\'emploi', JobOffer.query.count()],
        ['Offres actives', JobOffer.query.filter_by(is_active=True).count()]
    ]
    
    for row_idx, row_data in enumerate(stats_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_stats.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:  # En-têtes
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Candidatures
    ws_apps = wb.create_sheet("Candidatures")
    recent_applications = Application.query.order_by(Application.created_at.desc()).limit(10).all()
    
    headers = ['ID', 'Candidat', 'Poste', 'Service', 'Statut', 'Date']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_apps.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row_idx, app in enumerate(recent_applications, start=2):
        ws_apps.cell(row=row_idx, column=1, value=app.id)
        ws_apps.cell(row=row_idx, column=2, value=f"{app.applicant.first_name} {app.applicant.last_name}")
        ws_apps.cell(row=row_idx, column=3, value=app.position)
        ws_apps.cell(row=row_idx, column=4, value=app.service_type)
        ws_apps.cell(row=row_idx, column=5, value=app.status)
        ws_apps.cell(row=row_idx, column=6, value=app.created_at.strftime('%d/%m/%Y %H:%M'))
    
    # Rendez-vous
    ws_apts = wb.create_sheet("Rendez-vous")
    recent_appointments = Appointment.query.order_by(Appointment.created_at.desc()).limit(10).all()
    
    headers = ['ID', 'Client', 'Service', 'Date', 'Heure', 'Statut']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_apts.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row_idx, apt in enumerate(recent_appointments, start=2):
        ws_apts.cell(row=row_idx, column=1, value=apt.id)
        ws_apts.cell(row=row_idx, column=2, value=f"{apt.user.first_name} {apt.user.last_name}")
        ws_apts.cell(row=row_idx, column=3, value=apt.service_type)
        ws_apts.cell(row=row_idx, column=4, value=apt.date.strftime('%d/%m/%Y'))
        ws_apts.cell(row=row_idx, column=5, value=apt.time.strftime('%H:%M'))
        ws_apts.cell(row=row_idx, column=6, value=apt.status)
    
    # Offres d'emploi
    ws_jobs = wb.create_sheet("Offres d'emploi")
    job_offers = JobOffer.query.all()
    
    headers = ['ID', 'Titre', 'Entreprise', 'Localisation', 'Type', 'Statut', 'Candidatures']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_jobs.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    for row_idx, job in enumerate(job_offers, start=2):
        ws_jobs.cell(row=row_idx, column=1, value=job.id)
        ws_jobs.cell(row=row_idx, column=2, value=job.title)
        ws_jobs.cell(row=row_idx, column=3, value=job.company)
        ws_jobs.cell(row=row_idx, column=4, value=job.location)
        ws_jobs.cell(row=row_idx, column=5, value=job.contract_type)
        ws_jobs.cell(row=row_idx, column=6, value='Active' if job.is_active else 'Inactive')
        ws_jobs.cell(row=row_idx, column=7, value=len(job.applications))
    
    # Ajuster la largeur des colonnes
    for ws in [ws_stats, ws_apps, ws_apts, ws_jobs]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarder en mémoire
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=rapport_monderh_{datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")}.xlsx'}
    )

def export_pdf():
    """Exporte les données en PDF"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        spaceBefore=20
    )
    
    # Titre
    title = Paragraph(f"Rapport MondeRH - {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Statistiques générales
    elements.append(Paragraph("Statistiques Générales", heading_style))
    
    stats_data = [
        ['Métrique', 'Valeur'],
        ['Total Utilisateurs', str(User.query.count())],
        ['Total Candidatures', str(Application.query.count())],
        ['Total Rendez-vous', str(Appointment.query.count())],
        ['Newsletters Actives', str(Newsletter.query.filter_by(is_active=True).count())],
        ['Offres d\'emploi', str(JobOffer.query.count())],
        ['Offres actives', str(JobOffer.query.filter_by(is_active=True).count())]
    ]
    
    stats_table = Table(stats_data, colWidths=[200, 100])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(stats_table)
    elements.append(Spacer(1, 20))
    
    # Candidatures récentes
    elements.append(Paragraph("Candidatures Récentes", heading_style))
    
    recent_applications = Application.query.order_by(Application.created_at.desc()).limit(10).all()
    apps_data = [['ID', 'Candidat', 'Poste', 'Service', 'Statut', 'Date']]
    for app in recent_applications:
        apps_data.append([
            str(app.id),
            f"{app.applicant.first_name} {app.applicant.last_name}",
            app.position,
            app.service_type,
            app.status,
            app.created_at.strftime('%d/%m/%Y %H:%M')
        ])
    
    apps_table = Table(apps_data, colWidths=[30, 80, 80, 60, 50, 60])
    apps_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8)
    ]))
    elements.append(apps_table)
    elements.append(Spacer(1, 20))
    
    # Rendez-vous récents
    elements.append(Paragraph("Rendez-vous Récents", heading_style))
    
    recent_appointments = Appointment.query.order_by(Appointment.created_at.desc()).limit(10).all()
    apts_data = [['ID', 'Client', 'Service', 'Date', 'Heure', 'Statut']]
    for apt in recent_appointments:
        apts_data.append([
            str(apt.id),
            f"{apt.user.first_name} {apt.user.last_name}",
            apt.service_type,
            apt.date.strftime('%d/%m/%Y'),
            apt.time.strftime('%H:%M'),
            apt.status
        ])
    
    apts_table = Table(apts_data, colWidths=[30, 80, 60, 60, 50, 50])
    apts_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8)
    ]))
    elements.append(apts_table)
    
    # Générer le PDF
    doc.build(elements)
    buffer.seek(0)
    
    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=rapport_monderh_{datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")}.pdf'}
    )

def export_png():
    """Exporte les données en PNG (graphiques)"""
    # Créer une figure avec plusieurs sous-graphiques
    fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(15, 10))
    fig.suptitle(f'Rapport MondeRH - {datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M")}', fontsize=16, fontweight='bold')
    
    # Graphique 1: Statistiques générales
    stats_labels = ['Utilisateurs', 'Candidatures', 'Rendez-vous', 'Newsletters']
    stats_values = [
        User.query.count(),
        Application.query.count(),
        Appointment.query.count(),
        Newsletter.query.filter_by(is_active=True).count()
    ]
    
    ax1.bar(stats_labels, stats_values, color=['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4'])
    ax1.set_title('Statistiques Générales')
    ax1.set_ylabel('Nombre')
    ax1.tick_params(axis='x', rotation=45)
    
    # Graphique 2: Statut des candidatures
    applications = Application.query.all()
    status_counts = {}
    for app in applications:
        status_counts[app.status] = status_counts.get(app.status, 0) + 1
    
    if status_counts:
        ax2.pie(status_counts.values(), labels=status_counts.keys(), autopct='%1.1f%%')
        ax2.set_title('Statut des Candidatures')
    
    # Graphique 3: Offres d'emploi par type
    job_offers = JobOffer.query.all()
    contract_counts = {}
    for job in job_offers:
        contract_counts[job.contract_type] = contract_counts.get(job.contract_type, 0) + 1
    
    if contract_counts:
        ax3.bar(contract_counts.keys(), contract_counts.values(), color=['#FFEAA7', '#DDA0DD', '#98D8C8'])
        ax3.set_title('Offres par Type de Contrat')
        ax3.set_ylabel('Nombre')
        ax3.tick_params(axis='x', rotation=45)
    
    # Graphique 4: Évolution mensuelle (simulation)
    months = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Jun']
    applications_monthly = [5, 8, 12, 15, 18, 22]  # Données simulées
    
    ax4.plot(months, applications_monthly, marker='o', linewidth=2, markersize=8)
    ax4.set_title('Évolution des Candidatures')
    ax4.set_ylabel('Nombre de candidatures')
    ax4.grid(True, alpha=0.3)
    
    plt.tight_layout()
    
    # Sauvegarder en mémoire
    buffer = io.BytesIO()
    plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
    buffer.seek(0)
    plt.close()
    
    return Response(
        buffer.getvalue(),
        mimetype='image/png',
        headers={'Content-Disposition': f'attachment; filename=rapport_monderh_{datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")}.png'}
    )

@app.route('/admin/settings', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_settings():
    # Récupérer ou créer les paramètres
    settings = SiteSettings.query.first()
    if not settings:
        settings = SiteSettings()
        db.session.add(settings)
        db.session.commit()
    
    form = SiteSettingsForm(obj=settings)
    
    if form.validate_on_submit():
        form.populate_obj(settings)
        settings.updated_at = datetime.now(timezone.utc)
        db.session.commit()
        flash('Paramètres mis à jour avec succès !', 'success')
        return redirect(url_for('admin_settings'))
    
    return render_template('admin/settings.html', form=form, settings=settings)

@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    page = request.args.get('page', 1, type=int)
    users = User.query.paginate(page=page, per_page=20, error_out=False)
    return render_template('admin/users.html', users=users)

@app.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_edit_user(user_id):
    user = User.query.get_or_404(user_id)
    form = UserEditForm()
    
    if form.validate_on_submit():
        # Vérifier si l'email est déjà utilisé par un autre utilisateur
        existing_user = User.query.filter_by(email=form.email.data).first()
        if existing_user and existing_user.id != user.id:
            flash('Cette adresse email est déjà utilisée par un autre utilisateur.', 'error')
            return render_template('admin/edit_user.html', form=form, user=user)
        
        # Mettre à jour les informations de l'utilisateur
        user.first_name = form.first_name.data
        user.last_name = form.last_name.data
        user.email = form.email.data
        user.user_type = form.user_type.data
        user.company = form.company.data
        user.phone = form.phone.data
        user.is_active = form.is_active.data
        
        db.session.commit()
        flash('Utilisateur modifié avec succès !', 'success')
        return redirect(url_for('admin_users'))
    
    elif request.method == 'GET':
        # Pré-remplir le formulaire avec les données actuelles
        form.first_name.data = user.first_name
        form.last_name.data = user.last_name
        form.email.data = user.email
        form.user_type.data = user.user_type
        form.company.data = user.company
        form.phone.data = user.phone
        form.is_active.data = user.is_active
    
    return render_template('admin/edit_user.html', form=form, user=user)

@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_user(user_id):
    user = User.query.get_or_404(user_id)
    
    # Empêcher la suppression de son propre compte
    if user.id == current_user.id:
        flash('Vous ne pouvez pas supprimer votre propre compte.', 'error')
        return redirect(url_for('admin_users'))
    
    # Supprimer les données associées
    Application.query.filter_by(user_id=user.id).delete()
    Appointment.query.filter_by(user_id=user.id).delete()
    GoogleToken.query.filter_by(user_id=user.id).delete()
    
    # Supprimer l'utilisateur
    db.session.delete(user)
    db.session.commit()
    
    flash('Utilisateur supprimé avec succès !', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/applications')
@login_required
@admin_required
def admin_applications():
    page = request.args.get('page', 1, type=int)
    status_filter = request.args.get('status', '')
    service_filter = request.args.get('service', '')
    search_term = request.args.get('search', '')
    
    # Construire la requête de base
    query = Application.query
    
    # Appliquer les filtres
    if status_filter:
        query = query.filter(Application.status == status_filter)
    
    if service_filter:
        query = query.filter(Application.service_type == service_filter)
    
    # Appliquer la recherche
    if search_term:
        search_filter = f"%{search_term}%"
        query = query.filter(
            db.or_(
                Application.position.ilike(search_filter),
                Application.cover_letter.ilike(search_filter),
                User.first_name.ilike(search_filter),
                User.last_name.ilike(search_filter),
                User.email.ilike(search_filter)
            )
        ).join(User, Application.user_id == User.id)
    
    # Trier par date de création (plus récent en premier)
    applications = query.order_by(Application.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    
    return render_template('admin/applications.html', 
                         applications=applications,
                         status_filter=status_filter,
                         service_filter=service_filter,
                         search_term=search_term)

@app.route('/admin/applications/<int:application_id>')
@login_required
@admin_required
def admin_application_detail(application_id):
    application = Application.query.get_or_404(application_id)
    return render_template('admin/application_detail.html', application=application)

@app.route('/admin/applications/<int:application_id>/status', methods=['POST'])
@login_required
@admin_required
def admin_update_application_status(application_id):
    application = Application.query.get_or_404(application_id)
    status = request.form.get('status')
    
    if status in ['pending', 'reviewed', 'accepted', 'rejected']:
        application.status = status
        application.updated_at = datetime.now(timezone.utc)
        
        # Ajouter des notes si fournies
        notes = request.form.get('notes', '')
        if notes:
            application.notes = notes
        
        db.session.commit()
        
        # Envoyer un email de notification au candidat
        if application.applicant and application.applicant.email:
            try:
                if status == 'accepted':
                    send_application_accepted_email(application)
                elif status == 'rejected':
                    send_application_rejected_email(application)
                elif status == 'reviewed':
                    send_application_reviewed_email(application)
            except Exception as e:
                print(f"Erreur lors de l'envoi de l'email: {e}")
        
        flash(f'Statut de la candidature mis à jour avec succès !', 'success')
    else:
        flash('Statut invalide !', 'error')
    
    return redirect(url_for('admin_applications'))

@app.route('/admin/applications/<int:application_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_edit_application(application_id):
    application = Application.query.get_or_404(application_id)
    
    if request.method == 'POST':
        # Mettre à jour les informations de la candidature
        application.position = request.form.get('position', application.position)
        application.service_type = request.form.get('service_type', application.service_type)
        application.experience_years = request.form.get('experience_years', application.experience_years)
        application.salary_expectation = request.form.get('salary_expectation', application.salary_expectation)
        application.availability = request.form.get('availability', application.availability)
        application.cover_letter = request.form.get('cover_letter', application.cover_letter)
        application.notes = request.form.get('notes', application.notes)
        application.updated_at = datetime.now(timezone.utc)
        
        db.session.commit()
        flash('Candidature mise à jour avec succès !', 'success')
        return redirect(url_for('admin_application_detail', application_id=application.id))
    
    return render_template('admin/edit_application.html', application=application)

@app.route('/admin/applications/<int:application_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_application(application_id):
    application = Application.query.get_or_404(application_id)
    
    # Supprimer le fichier CV s'il existe
    if application.cv_filename:
        try:
            cv_path = os.path.join(app.config['UPLOAD_FOLDER'], application.cv_filename)
            if os.path.exists(cv_path):
                os.remove(cv_path)
        except Exception as e:
            print(f"Erreur lors de la suppression du fichier CV: {e}")
    
    db.session.delete(application)
    db.session.commit()
    
    flash('Candidature supprimée avec succès !', 'success')
    return redirect(url_for('admin_applications'))

@app.route('/admin/applications/export')
@login_required
@admin_required
def admin_export_applications():
    """Exporter les candidatures en CSV"""
    applications = Application.query.order_by(Application.created_at.desc()).all()
    
    # Créer le contenu CSV
    csv_data = []
    csv_data.append(['ID', 'Candidat', 'Email', 'Poste', 'Service', 'Expérience', 'Salaire', 'Disponibilité', 'Statut', 'Date'])
    
    for app in applications:
        candidate_name = f"{app.applicant.first_name} {app.applicant.last_name}" if app.applicant else "Candidat externe"
        candidate_email = app.applicant.email if app.applicant else "N/A"
        
        csv_data.append([
            app.id,
            candidate_name,
            candidate_email,
            app.position,
            app.service_type,
            app.experience_years,
            app.salary_expectation,
            app.availability,
            app.status,
            app.created_at.strftime('%d/%m/%Y')
        ])
    
    # Créer la réponse CSV
    output = StringIO()
    writer = csv.writer(output)
    writer.writerows(csv_data)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = 'attachment; filename=candidatures.csv'
    
    return response

@app.route('/admin/appointments')
@login_required
@admin_required
def admin_appointments():
    page = request.args.get('page', 1, type=int)
    appointments = Appointment.query.order_by(Appointment.date.desc()).paginate(page=page, per_page=20, error_out=False)
    return render_template('admin/appointments.html', appointments=appointments)

# Routes pour la gestion des offres d'emploi
@app.route('/admin/jobs')
@login_required
@admin_required
def admin_jobs():
    page = request.args.get('page', 1, type=int)
    jobs = JobOffer.query.order_by(JobOffer.created_at.desc()).paginate(page=page, per_page=20, error_out=False)
    return render_template('admin/jobs.html', jobs=jobs)

@app.route('/admin/jobs/new', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_new_job():
    form = JobOfferForm()
    if form.validate_on_submit():
        job = JobOffer(
            title=form.title.data,
            company=form.company.data,
            location=form.location.data,
            contract_type=form.contract_type.data,
            experience_level=form.experience_level.data,
            salary_range=form.salary_range.data,
            description=form.description.data,
            requirements=form.requirements.data,
            benefits=form.benefits.data,
            department=form.department.data,
            is_active=form.is_active.data
        )
        db.session.add(job)
        db.session.commit()
        flash('Offre d\'emploi créée avec succès !', 'success')
        return redirect(url_for('admin_jobs'))
    
    return render_template('admin/job_form.html', form=form, title="Nouvelle offre d'emploi")

@app.route('/admin/jobs/<int:job_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_edit_job(job_id):
    job = JobOffer.query.get_or_404(job_id)
    form = JobOfferForm(obj=job)
    
    if form.validate_on_submit():
        job.title = form.title.data
        job.company = form.company.data
        job.location = form.location.data
        job.contract_type = form.contract_type.data
        job.experience_level = form.experience_level.data
        job.salary_range = form.salary_range.data
        job.description = form.description.data
        job.requirements = form.requirements.data
        job.benefits = form.benefits.data
        job.department = form.department.data
        job.is_active = form.is_active.data
        job.updated_at = datetime.now(timezone.utc)
        
        db.session.commit()
        flash('Offre d\'emploi mise à jour avec succès !', 'success')
        return redirect(url_for('admin_jobs'))
    
    return render_template('admin/job_form.html', form=form, job=job, title="Modifier l'offre d'emploi")

@app.route('/admin/jobs/<int:job_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_job(job_id):
    job = JobOffer.query.get_or_404(job_id)
    db.session.delete(job)
    db.session.commit()
    flash('Offre d\'emploi supprimée avec succès !', 'success')
    return redirect(url_for('admin_jobs'))

@app.route('/admin/jobs/<int:job_id>/toggle', methods=['POST'])
@login_required
@admin_required
def admin_toggle_job(job_id):
    job = JobOffer.query.get_or_404(job_id)
    job.is_active = not job.is_active
    db.session.commit()
    status = "activée" if job.is_active else "désactivée"
    flash(f'Offre d\'emploi {status} avec succès !', 'success')
    return redirect(url_for('admin_jobs'))

# Route publique pour afficher les offres d'emploi
@app.route('/careers')
def careers():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    contract_type = request.args.get('contract_type', '')
    experience_level = request.args.get('experience_level', '')
    
    # Construire la requête avec filtres
    query = JobOffer.query.filter_by(is_active=True)
    
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            db.or_(
                JobOffer.title.ilike(search_term),
                JobOffer.company.ilike(search_term),
                JobOffer.location.ilike(search_term),
                JobOffer.description.ilike(search_term)
            )
        )
    
    if contract_type:
        query = query.filter(JobOffer.contract_type == contract_type)
    
    if experience_level:
        query = query.filter(JobOffer.experience_level == experience_level)
    
    jobs = query.order_by(JobOffer.created_at.desc()).paginate(page=page, per_page=12, error_out=False)
    return render_template('careers.html', jobs=jobs)

@app.route('/careers/<int:job_id>')
def job_detail(job_id):
    job = JobOffer.query.get_or_404(job_id)
    if not job.is_active:
        abort(404)
    
    # Récupérer les offres similaires (même département ou type de contrat)
    similar_jobs = JobOffer.query.filter(
        JobOffer.is_active == True,
        JobOffer.id != job.id
    ).filter(
        (JobOffer.department == job.department) | 
        (JobOffer.contract_type == job.contract_type) |
        (JobOffer.experience_level == job.experience_level)
    ).limit(3).all()
    
    # Si pas assez d'offres similaires, ajouter d'autres offres actives
    if len(similar_jobs) < 3:
        additional_jobs = JobOffer.query.filter(
            JobOffer.is_active == True,
            JobOffer.id != job.id,
            ~JobOffer.id.in_([j.id for j in similar_jobs])
        ).limit(3 - len(similar_jobs)).all()
        similar_jobs.extend(additional_jobs)
    
    return render_template('job_detail.html', job=job, similar_jobs=similar_jobs)

def get_google_callback_url():
    """Génère l'URL de callback Google en utilisant localhost"""
    callback_url = url_for('google_callback', _external=True)
    # Remplacer toutes les IPs privées par localhost
    if '10.188.193.170' in callback_url or '127.0.0.1' in callback_url:
        callback_url = callback_url.replace('10.188.193.170', 'localhost').replace('127.0.0.1', 'localhost')
    return callback_url

# Google OAuth routes
@app.route('/auth/google')
@login_required
@admin_required
def google_auth():
    """Initie l'authentification Google OAuth"""
    if not app.config['GOOGLE_CLIENT_ID'] or not app.config['GOOGLE_CLIENT_SECRET']:
        flash('Configuration Google OAuth manquante. Veuillez configurer GOOGLE_CLIENT_ID et GOOGLE_CLIENT_SECRET dans le fichier .env', 'error')
        return redirect(url_for('dashboard'))
    
    # Scopes nécessaires pour Drive et Calendar
    scopes = [
        'https://www.googleapis.com/auth/drive.file',
        'https://www.googleapis.com/auth/calendar'
    ]
    
    callback_url = get_google_callback_url()
    
    flow = Flow.from_client_config(
        {
            "web": {
                "client_id": app.config['GOOGLE_CLIENT_ID'],
                "client_secret": app.config['GOOGLE_CLIENT_SECRET'],
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [callback_url]
            }
        },
        scopes=scopes
    )
    
    flow.redirect_uri = callback_url
    
    # Générer un état sécurisé avec l'ID utilisateur
    import secrets
    state = secrets.token_urlsafe(32)
    
    authorization_url, _ = flow.authorization_url(
        access_type='offline',
        include_granted_scopes='true',
        state=state
    )
    
    # Stocker l'état dans la session avec l'ID utilisateur
    session['state'] = state
    session['google_auth_user_id'] = current_user.id
    
    return redirect(authorization_url)

@app.route('/auth/google/callback')
@login_required
@admin_required
def google_callback():
    """Callback après authentification Google"""
    # Vérifier que l'état est présent dans la session
    if 'state' not in session:
        flash('Erreur d\'authentification Google: état manquant', 'error')
        return redirect(url_for('dashboard'))
    
    # Vérifier que l'état dans l'URL correspond à celui de la session
    state_from_url = request.args.get('state')
    if not state_from_url or state_from_url != session['state']:
        flash('Erreur d\'authentification Google: état invalide', 'error')
        # Nettoyer la session
        session.pop('state', None)
        session.pop('google_auth_user_id', None)
        return redirect(url_for('dashboard'))
    
    # Vérifier que l'utilisateur correspond
    if 'google_auth_user_id' not in session or session['google_auth_user_id'] != current_user.id:
        flash('Erreur d\'authentification Google: utilisateur invalide', 'error')
        # Nettoyer la session
        session.pop('state', None)
        session.pop('google_auth_user_id', None)
        return redirect(url_for('dashboard'))
    
    callback_url = get_google_callback_url()
    
    flow = Flow.from_client_config(
        {
            "web": {
                "client_id": app.config['GOOGLE_CLIENT_ID'],
                "client_secret": app.config['GOOGLE_CLIENT_SECRET'],
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [callback_url]
            }
        },
        scopes=[
            'https://www.googleapis.com/auth/drive.file',
            'https://www.googleapis.com/auth/calendar'
        ],
        state=session['state']
    )
    
    flow.redirect_uri = callback_url
    
    try:
        # Pour le développement, permettre HTTP
        import os
        os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
        
        flow.fetch_token(authorization_response=request.url)
        credentials = flow.credentials
        
        # Sauvegarder les credentials
        save_google_credentials(current_user.id, credentials)
        
        # Vérifier que le token a été sauvegardé
        token = GoogleToken.query.filter_by(user_id=current_user.id).first()
        if token:
            flash('Connexion Google réussie ! Vous pouvez maintenant utiliser Google Drive et Calendar.', 'success')
        else:
            flash('Connexion réussie mais erreur lors de la sauvegarde du token.', 'warning')
        
    except Exception as e:
        print(f"Erreur Google callback: {e}")
        flash(f'Erreur lors de l\'authentification Google: {str(e)}', 'error')
    finally:
        # Nettoyer l'état de la session
        session.pop('state', None)
        session.pop('google_auth_user_id', None)
    
    return redirect(url_for('dashboard'))

@app.route('/auth/google/disconnect')
@login_required
@admin_required
def google_disconnect():
    """Déconnecte l'utilisateur de Google"""
    # Nettoyer la session
    session.pop('state', None)
    session.pop('google_auth_user_id', None)
    
    token = GoogleToken.query.filter_by(user_id=current_user.id).first()
    if token:
        db.session.delete(token)
        db.session.commit()
        flash('Déconnexion Google réussie !', 'success')
    else:
        flash('Aucune connexion Google active.', 'info')
    
    return redirect(url_for('dashboard'))

@app.route('/auth/google/clear-session')
@login_required
@admin_required
def google_clear_session():
    """Nettoie la session Google en cas de problème"""
    session.pop('state', None)
    session.pop('google_auth_user_id', None)
    flash('Session Google nettoyée. Vous pouvez réessayer l\'authentification.', 'info')
    return redirect(url_for('dashboard'))

@app.route('/auth/google/test')
@login_required
@admin_required
def google_test_config():
    """Test de la configuration Google OAuth (admin seulement)"""
    config_status = {
        'env_file': os.path.exists('.env'),
        'client_id': bool(app.config['GOOGLE_CLIENT_ID']),
        'client_secret': bool(app.config['GOOGLE_CLIENT_SECRET']),
        'client_id_format': app.config['GOOGLE_CLIENT_ID'].endswith('.apps.googleusercontent.com') if app.config['GOOGLE_CLIENT_ID'] else False
    }
    
    return jsonify({
        'status': 'success',
        'config': config_status,
        'message': 'Configuration Google OAuth vérifiée'
    })

@app.route('/api/google/status')
@login_required
@admin_required
def google_status():
    """Vérifier le statut de connexion Google de l'utilisateur"""
    token = GoogleToken.query.filter_by(user_id=current_user.id).first()
    
    if token:
        # Vérifier si le token n'est pas expiré
        if token.expiry and token.expiry < datetime.now():
            # Token expiré, le supprimer
            db.session.delete(token)
            db.session.commit()
            return jsonify({
                'connected': False,
                'message': 'Token expiré'
            })
        
        return jsonify({
            'connected': True,
            'message': 'Connecté à Google Workspace'
        })
    else:
        return jsonify({
            'connected': False,
            'message': 'Non connecté'
        })

@app.route('/google/drive')
@login_required
@admin_required
def google_drive_redirect():
    """Redirection vers Google Drive"""
    return redirect('https://drive.google.com')

@app.route('/google/calendar')
@login_required
@admin_required
def google_calendar_redirect():
    """Redirection vers Google Calendar"""
    return redirect('https://calendar.google.com')

@app.route('/google/forms')
@login_required
@admin_required
def google_forms_redirect():
    """Redirection vers Google Forms"""
    return redirect('https://forms.google.com')

@app.route('/google/sheets')
@login_required
@admin_required
def google_sheets_redirect():
    """Redirection vers Google Sheets"""
    return redirect('https://sheets.google.com')

# Create database tables
with app.app_context():
    try:
        db.create_all()
    except Exception as e:
        print(f"Note: Certaines tables existent déjà - {e}")

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000) 